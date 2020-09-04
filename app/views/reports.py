from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404, JsonResponse
from django.views import View
from collections import OrderedDict, defaultdict
from django.db.models import *
from app.models import *
from app.forms import *
from app.helpers import *
from app.other_constants import user_constants
from django.utils import timezone, safestring
from django.contrib.auth.models import User

from django.conf import settings

# import datetime
from datetime import datetime

import os
import calendar

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THICK, BORDER_THIN
from openpyxl.styles.fills import Fill, PatternFill 
from openpyxl.styles.fonts import Font

from easy_pdf.rendering import render_to_pdf_response


def get_reports_pdf(request):
    template_name = 'app/app_files/reports/pdf_report.html'
    context = {'html_data': request.POST["html_content"]} 
    return render_to_pdf_response(request,template_name,context)



class GSTLedgerReportsView(View):
    # Template 
    template_name = 'app/app_files/reports/index.html'

    # Initialize 
    data = defaultdict()
    data["active_link"] = 'GST Report'
    data["breadcrumb_title"] = 'GST Ledger Reports'

    # Custom CSS/JS Files For Inclusion into template
    data["css_files"] = []
    data["js_files"] = ['custom_files/js/reports.js',]

    data["included_template"] = 'app/app_files/reports/gst_ledger_reports.html'

    #****************************************************************************
    #
    #****************************************************************************
    def create_xls(self, query_dict = None, filename="reports.xlsx", headers=[], personal_data = None, time_period = False, start_date = "", end_date = ""):

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reports"

        thick_border = Border(left=Side(style=BORDER_THICK), 
                     right=Side(style=BORDER_THICK), 
                     top=Side(style=BORDER_THICK), 
                     bottom=Side(style=BORDER_THICK))

        thin_border = Border(left=Side(style=BORDER_THIN), 
                     right=Side(style=BORDER_THIN), 
                     top=Side(style=BORDER_THIN), 
                     bottom=Side(style=BORDER_THIN))

        header_fill = PatternFill(fgColor='363231', fill_type = 'solid')

        header_font = Font(size=12, color='FFFFFF', bold=True)

        if personal_data is not None:
            ws.cell(row=1, column=1).value = "Name"
            ws.cell(row=1, column=2).value = personal_data[0]

            ws.cell(row=2, column=1).value = "Organisation Name"
            ws.cell(row=2, column=2).value = personal_data[1]

            ws.cell(row=3, column=1).value = "Address"
            ws.cell(row=3, column=2).value = ""

            ws.cell(row=4, column=1).value = "GST Number"
            ws.cell(row=4, column=2).value = ""

            ws.cell(row=5, column=1).value = "Ledger Generation Date"
            ws.cell(row=5, column=2).value = timezone.now().strftime("%Y-%m-%d")

            #
            #
            if time_period is False:

                ws.cell(row=6, column=1).value = "Start Date"
                ws.cell(row=6, column=2).value = start_date

                ws.cell(row=7, column=1).value = "End Date"
                ws.cell(row=7, column=2).value = end_date

            else:
                ws.cell(row=6, column=1).value = "Report Type"

                time_dict = {'1':"Monthly", '2': "Quaterly", '3':"Half Yearly", '4': "Yearly"}
                ws.cell(row=6, column=2).value = time_dict[time_period]


        for head in range(1, len(headers)+1):
            ws.cell(row=9, column=head).value = headers[head-1]
            ws.cell(row=9, column=head).border = thick_border
            ws.cell(row=9, column=head).fill = header_fill
            ws.cell(row=9, column=head).font = header_font

        index_col = None
        if "Month" in headers:
            index_col = headers.index("Month")

        date_col = None
        if "Created On" in headers:
            date_col = headers.index("Created On")

        cal_colm = 0
        if "Credit" in headers:
            cal_colm = headers.index("Credit")
        if "Debit" in headers:
            cal_colm = headers.index("Debit")    

        col = 1
        row = 10

        for row in range(10, len(query_dict)+10):
            for col in range(1, len(query_dict[row-10])+1):


                if index_col is not None and col == index_col+1:                  
                    ws.cell(row=row, column=col).value = calendar.month_name[query_dict[row-10][col-1]]
                elif date_col is not None and col == date_col+1:
                    ws.cell(row=row, column=col).value = query_dict[row-10][col-1].strftime("%Y-%m-%d")
                else:
                    if col == cal_colm+1:
                        ws.cell(row=row, column=col).value = float(query_dict[row-10][col-1])
                    else:
                        ws.cell(row=row, column=col).value = query_dict[row-10][col-1]
        
                ws.cell(row=row, column=col).border = thin_border
                
        chrr = chr(col+65-1)
        head_crr = chr(col+65-2)

        try:
            ws[head_crr+str(row+1)] = "TOTAL"
            ws[head_crr+str(row+1)].border = thick_border
            ws[head_crr+str(row+1)].fill = header_fill
            ws[head_crr+str(row+1)].font = header_font

            ws[chrr+str(row+1)] = '= SUM('+chrr+'10:'+chrr+str(row)+')'
            ws[chrr+str(row+1)].border = thick_border
            ws[chrr+str(row+1)].fill = header_fill
            ws[chrr+str(row+1)].font = header_font

        except:
            pass

        try:
            
            wb.save(os.path.join(settings.REPORTS,filename))
        except:
            pass

    
    #****************************************************************************
    #
    #****************************************************************************
    def get(self, request, *args, **kwargs):

        save_btn = request.GET.get("save_btn", False)
        #
        #

        user = User.objects.get(pk = request.user.id)
        organisation = users_model.Organisations.objects.get(user = user)
        name = user.username
        organisation_name = organisation.organisation_name

        try:
            organ_addr = users_model.User_Address_Details.objects.get(
                is_organisation = True, default_address = True, organisation = organisation
            )
            self.data["organ_addr"] = organ_addr
            if(organ_addr.organisation_tax is not None):
                self.data["gst_number"] = organ_addr.organisation_tax.gstin
        except:
            organ_addr = users_model.User_Address_Details.objects.filter(
                is_organisation = True, organisation = organisation
            )
            self.data["organ_addr"] = organ_addr[0]
            if(organ_addr[0].organisation_tax is not None):
                self.data["gst_number"] = organ_addr[0].organisation_tax.gstin
            # self.data["address"] = ""

        self.data["user_name"] = user
        self.data["organisation_name"] = organisation_name
    
        # self.data["gst_number"] = "aaaa"
        self.data["ledger_date"] = timezone.now().strftime("%Y-%m-%d")

        personal_data = [name, organisation_name]  

        year_xx = timezone.now().year
        self.data["year_list"] = [year for year in range(year_xx - 10, year_xx+1)]

        self.data["gst_reports_show"] = False
        self.data["pdf_btn"] = False

        #self.data["gst_numbers_list"] = gst_ledger_model.GST_Ledger.objects.filter(user = request.user).distinct('gst_number').values_list('gst_number')

        #print(self.data["gst_numbers_list"])


        if save_btn:

            self.data["gst_reports_show"] = True
            self.data["pdf_btn"] = True

            start_date = request.GET.get("start_date", "")
            # change start date formate
            # start_date = datetime.strptime(str(s_date), '%d-%m-%Y').strftime('%Y-%m-%d')
            end_date = request.GET.get("end_date", "")
            # change end date formate
            # end_date = datetime.strptime(str(e_date), '%d-%m-%Y').strftime('%Y-%m-%d')
            account_type = request.GET.get("account_type", None)
            time_period = request.GET.get("time_period", False)
            year_t = request.GET.get("year", None)
            month_t = request.GET.get("month", None)
            month_tq = request.GET.get("q_month", None)
            month_th = request.GET.get("h_month", None)

            self.data["start_date"] = start_date
            self.data["end_date"] = end_date
            self.data["account_type"] = account_type
            self.data["time_period"] = time_period
            self.data["year_t"] = year_t
            self.data["month_t"] = month_t
            self.data["month_tq"] = month_tq
            self.data["month_th"] = month_th

            filename = "Reports.xlsx"

            if time_period == '0':
                time_period = False


            if month_tq:                
                month_tq_l = [(int(month_tq) * 3) + i for i in range(1,4)] 

            if month_th:
                month_th_l = {'1':[4,5,6,7,8,9],'2':[1,2,3,10,11,12]}

              

            g_type = None
            #
            #
            gst_reports = gst_ledger_model.GST_Ledger.objects

            if time_period is False and (start_date !="" and start_date!= False):
                #print("run start")
                gst_reports = gst_reports.filter(created_on__gte = start_date)

            if time_period is False and (end_date !="" and end_date!= False):
                #print("run end")
                gst_reports = gst_reports.filter(created_on__lte = end_date)


            all_0 = True if account_type == '0' else False
            all_1 = True if account_type == '1' else False

            self.data["total_idt"] = True if account_type == '0' else False

            #
            #
            if time_period and all_0:

                gst_reports = gst_reports.filter(input_tab = True)

                #
                # Monthly ALL 0
                if time_period == '1' and month_t is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month = month_t) \
                            .order_by('created_on__month')

                    filename = "Output_Monthly.xlsx"

                # Quaterly -ALL 0
                if time_period == '2' and month_tq is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_tq_l) \
                            .order_by('created_on__month')
                    filename = "Output_Quaterly.xlsx"

                # Half Yearly -ALL 0
                if time_period == '3' and month_th is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_th_l[month_th]) \
                            .order_by('created_on__month')
                    filename = "Output_Quaterly.xlsx"                                                                

                # Yearly - ALL 0
                if time_period == '4':
                    gst_reports = gst_reports.filter(created_on__year = year_t).order_by('created_on__month')
                    filename = "Output_Yearly.xlsx"

                if len(gst_reports) > 0:  

                    # xls_reports = gst_reports.values_list('is_purchase_order','is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','purchase_order__vendor__contact_name','purchase_order__purchase_order_number','total_tax')
                    # xls_reports = gst_reports.values_list('is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','total_tax')
                    xls_reports = gst_reports.values_list('is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','total_tax','is_expense','expense__vendor__contact_name','expense__exp_number')
                    # print(xls_reports[])
                
                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = ['','','','','']

                        if xls_reports[i][5]:
                            new_data[3] = "EXPENSE"
                            new_data[1] = xls_reports[i][6]
                            new_data[2] = xls_reports[i][7]
                        if xls_reports[i][0]:
                            new_data[3] = "CREDITNOTE"
                            new_data[1] = xls_reports[i][2]
                            new_data[2] = xls_reports[i][3]
                        
                        new_data[0] = xls_reports[i][1]
                        new_data[4] = xls_reports[i][4]

                        main.append(new_data)


                    headers = ["Created On","Particulars","Voucher","Voucher Type","Debit"]
                    self.create_xls(
                        main,
                        filename,
                        headers,
                        personal_data,
                        time_period,
                        self.data["start_date"],
                        self.data["end_date"],
                    )

            #
            #
            elif time_period and all_1:

                gst_reports = gst_reports.filter(input_tab = False)
                #
                # Monthly - ALL
                if time_period == '1' and month_t is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month = month_t) \
                            .order_by('created_on__month')
                    filename = "All_Input_Monthly.xlsx"

                # Quaterly - ALL 1
                if time_period == '2' and month_tq is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_tq_l) \
                            .order_by('created_on__month')
                    filename = "Input_Quaterly.xlsx"

                # Half Yearly -ALL 0
                if time_period == '3' and month_th is not None:
                    gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_th_l[month_th]) \
                            .order_by('created_on__month')
                    filename = "Input_Quaterly.xlsx" 

                # Yearly - ALL 1
                if time_period == '4':
                    gst_reports = gst_reports.filter(created_on__year = year_xx).order_by('created_on__month') 
                    filename = "All_Input_Yearly.xlsx"
                    
                if len(gst_reports) > 0:  
                    xls_reports = gst_reports.values_list('created_on','invoice__invoice_customer__contact_name','invoice__invoice_number','total_tax')

                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = []

                        for x in range(0,len(xls_reports[i])):
                            if x == 3:
                                new_data.append('INVOICE')
                                new_data.append(xls_reports[i][x])
                            else:          
                                new_data.append(xls_reports[i][x])
                        main.append(new_data)

                    headers = ["Created On","Particulars","Voucher","Voucher Type","Credit"]
                    self.create_xls(
                        main,
                        filename,
                        headers,
                        personal_data,
                        time_period,
                        self.data["start_date"],
                        self.data["end_date"],
                    )

            elif not time_period or time_period == '0':
                
                if all_0:

                    gst_reports = gst_reports.filter(input_tab = True) 

                    #xls_reports = gst_reports.values_list('is_purchase_order','is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','purchase_order__vendor__contact_name','purchase_order__purchase_order_number','total_tax')
                    # xls_reports = gst_reports.values_list('is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','total_tax')
                    xls_reports = gst_reports.values_list('is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','total_tax','is_expense','expense__vendor__contact_name','expense__exp_number')

                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = ['','','','','']

                        if xls_reports[i][5]:
                            new_data[3] = "EXPENSE"
                            new_data[1] = xls_reports[i][6]
                            new_data[2] = xls_reports[i][7]
                        if xls_reports[i][0]:
                            new_data[3] = "CREDITNOTE"
                            new_data[1] = xls_reports[i][2]
                            new_data[2] = xls_reports[i][3]
                        
                        new_data[0] = xls_reports[i][1]
                        new_data[4] = xls_reports[i][4]

                        """if xls_reports[i][0]:
                            new_data[3] = "CREDITNOTE"
                            new_data[1] = xls_reports[i][2]
                            new_data[2] = xls_reports[i][3]
                        
                        new_data[0] = xls_reports[i][1]
                        new_data[4] = xls_reports[i][4]"""

                        main.append(new_data)

                    headers = ["Created On","Particulars","Voucher","Voucher Type","Credit"]
                    self.create_xls(
                        main,
                        "All_Output.xlsx",
                        headers,
                        personal_data
                    )

                elif all_1:
                    gst_reports = gst_reports.filter(input_tab = False)
                
                    xls_reports = gst_reports.values_list('created_on','invoice__invoice_customer__contact_name','invoice__invoice_number','total_tax')

                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = []

                        for x in range(0,len(xls_reports[i])):
                            if x == 3:
                                new_data.append('INVOICE')
                                new_data.append(xls_reports[i][x])
                            else:          
                                new_data.append(xls_reports[i][x])
                        main.append(new_data)

                    headers = ["Created On","Particulars","Voucher","Voucher Type","Debit"]
                    self.create_xls(
                        main,
                        "All_Input.xlsx",
                        headers,
                        personal_data,
                        time_period,
                        self.data["start_date"],
                        self.data["end_date"],
                    )
            else:

                splitr = account_type.split("-")

                self.data["total_idt"] = bool(splitr[0])

                g_type = splitr[1] 

                if splitr[0] == '0':                    

                    gst_reports = gst_reports.filter(input_tab = True) 

                    if splitr[1] == 'cgst':
                        gst_reports = gst_reports.filter(cgst_amount__gte = 1)                            
                        filename = "Output_CGST.xlsx"                

                    if splitr[1] == 'sgst':
                        gst_reports = gst_reports.filter(sgst_amount__gte = 1)
                        filename = "Output_SGST.xlsx"

                    if splitr[1] == 'igst':
                        gst_reports = gst_reports.filter(igst_amount__gte = 1)
                        filename = "Output_IGST.xlsx"

                    if time_period == '1':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month = month_t)  

                    if time_period == '2':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_tq)

                    if time_period == '3':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_th_l[month_th])

                    if time_period == '4':
                        gst_reports = gst_reports.filter(created_on__year = year_t)    

                    xls_reports = gst_reports.values_list('is_creditnote', 'created_on','creditnote__contact_name__contact_name','creditnote__credit_number','total_tax','is_expense','expense__vendor__contact_name','expense__exp_number')


                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = ['','','','','']

                        if xls_reports[i][5]:
                            new_data[3] = "EXPENSE"
                            new_data[1] = xls_reports[i][6]
                            new_data[2] = xls_reports[i][7]
                        if xls_reports[i][0]:
                            new_data[3] = "CREDITNOTE"
                            new_data[1] = xls_reports[i][2]
                            new_data[2] = xls_reports[i][3]
                        
                        new_data[0] = xls_reports[i][1]
                        new_data[4] = xls_reports[i][4]

                        main.append(new_data)  

                    headers = ["Created On","Particulars","Voucher","Voucher Type","Credit"]
                    self.create_xls(
                        main,
                        filename,
                        headers,
                        personal_data,
                        time_period,
                        self.data["start_date"],
                        self.data["end_date"],
                    )       

                if splitr[0] == '1': 

                    gst_reports = gst_reports.filter(input_tab = False) 

                    if splitr[1] == 'cgst':
                        gst_reports = gst_reports.filter(cgst_amount__gte = 1)
                        filename = "Input_CGST.xlsx"

                    if splitr[1] == 'sgst':
                        gst_reports = gst_reports.filter(sgst_amount__gte = 1)
                        filename = "Input_SGST.xlsx"

                    if splitr[1] == 'igst':
                        gst_reports = gst_reports.filter(igst_amount__gte = 1)
                        filename = "Input_IGST.xlsx"
                
                    if time_period == '1':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month = month_t)  

                    if time_period == '2':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_tq)

                    if time_period == '3':
                        gst_reports = gst_reports.filter(created_on__year = year_t, created_on__month__in = month_th_l[month_th])

                    if time_period == '4':
                        gst_reports = gst_reports.filter(created_on__year = year_t) 

                    xls_reports = gst_reports.values_list('created_on','invoice__invoice_customer__contact_name','invoice__invoice_number','total_tax')

                    main = []
                    for i in range(0,len(xls_reports)):
                        new_data = []

                        for x in range(0,len(xls_reports[i])):
                            if x == 3:
                                new_data.append('INVOICE')
                                new_data.append(xls_reports[i][x])
                            else:          
                                new_data.append(xls_reports[i][x])
                        main.append(new_data)

                    headers = ["Created On","Particulars","Voucher","Voucher Type","Debit"]
                    self.create_xls(
                        main,
                        filename,
                        headers,
                        personal_data,
                        time_period,
                        self.data["start_date"],
                        self.data["end_date"],
                    )

            self.data["g_type"] = g_type

            self.data["gst_reports"] = gst_reports

            #
            #
            main_tax = 0

            for row in gst_reports.values():
                if all_1 or all_0:
                    main_tax += float(row["total_tax"])
                else:
                    if g_type == 'cgst':
                        main_tax += float(row["cgst_amount"])

                    if g_type == 'sgst':
                        main_tax += float(row["sgst_amount"])

                    if g_type == 'igst':
                        main_tax += float(row["igst_amount"])

            self.data["main_tax"] = '{0:.2f}'.format(main_tax)
            self.data["filename"] = filename

            try:
                org = users_model.Organisations.objects.get(user = request.user)

                org_info = users_model.Organisation_Contact.objects.get(organisation = org)

                self.data["org_email"] = org_info.email
            except:
                self.data["org_email"] = ""

        return render(request, self.template_name, self.data)


#
#
#        
def send_email(request):
    if request.POST:
        
        subject = "Reports Mail"
        msg_body = 'test'

        msg_html = "<html><body>"+msg_body+"</body></html>"

        to_list = [request.POST["email_address"]]
        cc_list = []
    
        attachements = []
        attachements = [os.path.join(settings.REPORTS,request.POST["file_name"])]   
    
        msg = email_helper.Email_Helper(to=to_list, cc=cc_list, subject=subject, message=msg_html, attachment=attachements)
        msg.mail_send()

    return redirect("/reports/gst_ledger/")